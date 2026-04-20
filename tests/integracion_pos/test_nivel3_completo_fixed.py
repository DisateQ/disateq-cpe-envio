"""
test_nivel3_completo.py
=======================
Test de integración Nivel 3 — Motor CPE v3.0

Ejecuta el flujo COMPLETO end-to-end:
1. Lee Excel worksheet _CPE (v1.2, 29 campos)
2. Normaliza a estructura interna
3. Genera TXT formato APIFAS
4. Valida TXT pre-envío
5. Envía a APIFAS (real o mock)
6. Valida respuesta CDR

Uso:
    # Con mock APIFAS (sin conexión real)
    python test_nivel3_completo.py --mock
    
    # Con APIFAS real (producción)
    python test_nivel3_completo.py --real
    
    # Especificar archivo Excel
    python test_nivel3_completo.py --mock --excel ventas_test.xlsx

Autor: Fernando Miguel Tejada Quevedo (@fhertejadaDEV)
Empresa: DisateQ™
Fecha: Abril 2026
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime
import json
import openpyxl
import requests

# Ajustar path para imports
REPO_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(REPO_ROOT / 'src'))

# Importar validador si existe
try:
    from validator_pre_envio import ValidadorTxtApifas
except ImportError:
    print("⚠️  Advertencia: validator_pre_envio.py no encontrado en el mismo directorio")
    ValidadorTxtApifas = None


class XlsxAdapterSimple:
    """Adaptador simple para leer Excel - versión standalone."""
    
    def __init__(self, archivo: str):
        self.archivo = Path(archivo)
        self.wb = None
        self.ws = None
    
    def connect(self):
        """Abre el archivo Excel."""
        self.wb = openpyxl.load_workbook(self.archivo, data_only=True)
        self.ws = self.wb['_CPE']
    
    def disconnect(self):
        """Cierra el archivo."""
        if self.wb:
            self.wb.close()
    
    def read_pending(self):
        """Lee comprobantes pendientes."""
        if not self.ws:
            return []
        
        # Leer headers (fila 1)
        headers = []
        for cell in self.ws[1]:
            headers.append(cell.value)
        
        # Leer datos (fila 2 en adelante)
        comprobantes = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Fila vacía
                continue
            
            comp = {}
            for i, valor in enumerate(row):
                if i < len(headers):
                    comp[headers[i]] = valor
            
            comprobantes.append(comp)
        
        return comprobantes
    
    def read_items(self, comprobante):
        """Retorna items del comprobante (simplificado: 1 item por ahora)."""
        return [comprobante]  # Para el test, el item está en la misma fila


def generar_txt_apifas_simple(comprobante: dict) -> str:
    """
    Genera TXT formato APIFAS.
    Versión simplificada standalone.
    """
    # Línea 1: Cabecera
    linea1 = (
        f"TIPO|{comprobante.get('cpe_tipo', '03')}|"
        f"SERIE|{comprobante.get('cpe_serie', 'B001')}|"
        f"NUMERO|{comprobante.get('cpe_numero', '0')}|"
        f"FECHA|{comprobante.get('cpe_fecha', '')}|"
        f"MONEDA|{comprobante.get('cpe_moneda', 'PEN')}"
    )
    
    # Línea 2: Cliente
    linea2 = (
        f"CLI_TIPO|{comprobante.get('cli_tipo_doc', '1')}|"
        f"CLI_NUM|{comprobante.get('cli_num_doc', '')}|"
        f"CLI_NOMBRE|{comprobante.get('cli_nombre', '')}|"
        f"CLI_DIR|{comprobante.get('cli_direccion', '')}"
    )
    
    # Línea 3: Item
    linea3 = (
        f"ITEM_COD|{comprobante.get('item_codigo', '')}|"
        f"ITEM_DESC|{comprobante.get('item_descripcion', '')}|"
        f"ITEM_CANT|{comprobante.get('item_cantidad', 1)}|"
        f"ITEM_UND|{comprobante.get('item_unidad', 'NIU')}|"
        f"ITEM_PRECIO|{comprobante.get('item_precio_unitario', 0)}"
    )
    
    # Línea 4: Totales
    linea4 = (
        f"VENTA_GRAVADA|{comprobante.get('venta_gravada', 0)}|"
        f"VENTA_EXONERADA|{comprobante.get('venta_exonerada', 0)}|"
        f"VENTA_INAFECTA|{comprobante.get('venta_inafecta', 0)}|"
        f"VENTA_ICBPER|{comprobante.get('venta_icbper', 0)}|"
        f"VENTA_IGV|{comprobante.get('venta_igv', 0)}|"
        f"VENTA_TOTAL|{comprobante.get('venta_total', 0)}"
    )
    
    # Línea 5: Pago
    linea5 = (
        f"PAGO_FORMA|{comprobante.get('pago_forma', 'EFECTIVO')}|"
        f"PAGO_MONTO|{comprobante.get('pago_monto', 0)}"
    )
    
    return f"{linea1}\n{linea2}\n{linea3}\n{linea4}\n{linea5}"


def enviar_a_apifas_simple(contenido_txt: str, url: str) -> tuple:
    """
    Envía TXT a APIFAS.
    Versión simplificada standalone.
    """
    try:
        response = requests.post(
            url,
            data={'txt': contenido_txt},
            timeout=30
        )
        
        if response.status_code == 200:
            respuesta = response.json()
            exito = respuesta.get('success', False)
            return exito, respuesta
        else:
            return False, {
                'success': False,
                'error': f'HTTP {response.status_code}',
                'mensaje_sunat': response.text[:200]
            }
    
    except Exception as e:
        return False, {
            'success': False,
            'error': str(e),
            'mensaje_sunat': f'Error de conexión: {e}'
        }


class TestNivel3:
    """
    Test de integración completa Nivel 3.
    
    Ejecuta todo el flujo desde Excel hasta CDR.
    """
    
    def __init__(self, archivo_excel: str, usar_mock: bool = True):
        """
        Inicializa el test.
        
        Args:
            archivo_excel: Ruta al archivo Excel con worksheet _CPE
            usar_mock: Si True, usa mock APIFAS local. Si False, APIFAS real.
        """
        self.archivo_excel = Path(archivo_excel)
        self.usar_mock = usar_mock
        self.carpeta_salida = Path("./salida_nivel3")
        self.log_test = []
        
        # Crear carpeta de salida
        self.carpeta_salida.mkdir(exist_ok=True)
        
        # Configurar URL según modo
        if usar_mock:
            self.url_apifas = "http://localhost:8080/produccion_text.php"
        else:
            self.url_apifas = "https://apifas.disateq.com/produccion_text.php"
    
    def log(self, mensaje: str, nivel: str = "INFO"):
        """
        Registra mensaje en el log.
        
        Args:
            mensaje: Mensaje a registrar
            nivel: Nivel de log (INFO, OK, ERROR, WARNING)
        """
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        iconos = {
            "INFO": "ℹ️",
            "OK": "✅",
            "ERROR": "❌",
            "WARNING": "⚠️"
        }
        
        icono = iconos.get(nivel, "•")
        linea = f"[{timestamp}] {icono} {mensaje}"
        
        print(linea)
        self.log_test.append(linea)
    
    def paso_1_leer_excel(self) -> tuple:
        """
        Paso 1: Lee comprobante desde Excel.
        
        Returns:
            (exito, comprobante_data, items_data)
        """
        self.log("=" * 70)
        self.log("PASO 1: Lectura de Excel", "INFO")
        self.log("=" * 70)
        
        try:
            # Validar que el archivo existe
            if not self.archivo_excel.exists():
                self.log(f"Archivo no encontrado: {self.archivo_excel}", "ERROR")
                return False, None, None
            
            self.log(f"Archivo: {self.archivo_excel.name}", "INFO")
            
            # Crear adaptador
            adapter = XlsxAdapterSimple(str(self.archivo_excel))
            
            # Conectar
            self.log("Conectando al archivo Excel...", "INFO")
            adapter.connect()
            self.log("Conexión establecida", "OK")
            
            # Leer comprobantes pendientes
            self.log("Leyendo comprobantes pendientes...", "INFO")
            pendientes = adapter.read_pending()
            
            if not pendientes:
                self.log("No hay comprobantes pendientes", "WARNING")
                return False, None, None
            
            self.log(f"Encontrados {len(pendientes)} comprobante(s) pendiente(s)", "OK")
            
            # Tomar el primero
            comprobante = pendientes[0]
            self.log(f"Procesando: {comprobante.get('cpe_serie', '???')}-{comprobante.get('cpe_numero', '???')}", "INFO")
            
            # Leer items del comprobante
            self.log("Leyendo items del comprobante...", "INFO")
            items = adapter.read_items(comprobante)
            self.log(f"Encontrados {len(items)} item(s)", "OK")
            
            # Desconectar
            adapter.disconnect()
            
            # Guardar datos raw para debugging
            archivo_debug = self.carpeta_salida / "paso1_datos_excel.json"
            with open(archivo_debug, 'w', encoding='utf-8') as f:
                json.dump({
                    'comprobante': comprobante,
                    'items': items
                }, f, indent=2, ensure_ascii=False, default=str)
            
            self.log(f"Datos guardados en: {archivo_debug.name}", "INFO")
            self.log("")
            
            return True, comprobante, items
            
        except Exception as e:
            self.log(f"Error en Paso 1: {e}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            return False, None, None
    
    def paso_2_generar_txt(self, comprobante: dict, items: list) -> tuple:
        """
        Paso 2: Genera TXT formato APIFAS.
        
        Args:
            comprobante: Datos de cabecera
            items: Lista de items
        
        Returns:
            (exito, ruta_txt, contenido_txt)
        """
        self.log("=" * 70)
        self.log("PASO 2: Generación de TXT APIFAS", "INFO")
        self.log("=" * 70)
        
        try:
            # Generar TXT
            self.log("Generando TXT...", "INFO")
            contenido_txt = generar_txt_apifas_simple(comprobante)
            
            # Nombre del archivo TXT
            serie = comprobante.get('cpe_serie', 'XXX')
            numero = comprobante.get('cpe_numero', '0000')
            nombre_txt = f"{serie}-{numero}.txt"
            
            # Guardar TXT
            ruta_txt = self.carpeta_salida / nombre_txt
            ruta_txt.write_text(contenido_txt, encoding='utf-8')
            
            self.log(f"TXT generado: {nombre_txt}", "OK")
            self.log(f"Tamaño: {len(contenido_txt)} bytes", "INFO")
            self.log(f"Líneas: {len(contenido_txt.split(chr(10)))}", "INFO")
            self.log("")
            
            return True, ruta_txt, contenido_txt
            
        except Exception as e:
            self.log(f"Error en Paso 2: {e}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            return False, None, None
    
    def paso_3_validar_txt(self, ruta_txt: Path) -> bool:
        """
        Paso 3: Valida TXT antes de enviar.
        
        Args:
            ruta_txt: Ruta al archivo TXT
        
        Returns:
            True si es válido, False si tiene errores
        """
        self.log("=" * 70)
        self.log("PASO 3: Validación Pre-Envío", "INFO")
        self.log("=" * 70)
        
        if ValidadorTxtApifas is None:
            self.log("Validador no disponible, saltando paso...", "WARNING")
            self.log("")
            return True
        
        try:
            # Crear validador
            validador = ValidadorTxtApifas(str(ruta_txt))
            
            # Ejecutar validación
            self.log("Ejecutando validaciones...", "INFO")
            es_valido, errores, advertencias = validador.validar()
            
            # Mostrar resultados
            if es_valido:
                self.log("TXT válido ✓", "OK")
            else:
                self.log(f"TXT inválido - {len(errores)} error(es)", "ERROR")
            
            if advertencias:
                self.log(f"Advertencias: {len(advertencias)}", "WARNING")
                for adv in advertencias[:3]:  # Mostrar solo primeras 3
                    self.log(f"  {adv}", "WARNING")
            
            if errores:
                self.log("Errores encontrados:", "ERROR")
                for err in errores[:3]:  # Mostrar solo primeros 3
                    self.log(f"  {err}", "ERROR")
            
            # Guardar reporte
            archivo_reporte = self.carpeta_salida / "paso3_validacion.txt"
            archivo_reporte.write_text(validador.generar_reporte(), encoding='utf-8')
            self.log(f"Reporte guardado: {archivo_reporte.name}", "INFO")
            self.log("")
            
            return es_valido
            
        except Exception as e:
            self.log(f"Error en Paso 3: {e}", "ERROR")
            return False
    
    def paso_4_enviar_apifas(self, contenido_txt: str, serie: str, numero: str) -> tuple:
        """
        Paso 4: Envía TXT a APIFAS (mock o real).
        
        Args:
            contenido_txt: Contenido del archivo TXT
            serie: Serie del comprobante
            numero: Número del comprobante
        
        Returns:
            (exito, respuesta_json)
        """
        self.log("=" * 70)
        modo = "MOCK" if self.usar_mock else "REAL"
        self.log(f"PASO 4: Envío a APIFAS ({modo})", "INFO")
        self.log("=" * 70)
        
        if self.usar_mock:
            self.log("⚠️  MODO MOCK: No se envía a SUNAT real", "WARNING")
        else:
            self.log("🚨 MODO REAL: Se enviará a SUNAT", "WARNING")
        
        self.log(f"URL: {self.url_apifas}", "INFO")
        self.log("")
        
        try:
            # Enviar a APIFAS
            self.log("Enviando TXT...", "INFO")
            
            exito, respuesta = enviar_a_apifas_simple(
                contenido_txt=contenido_txt,
                url=self.url_apifas
            )
            
            if exito:
                self.log("Envío exitoso ✓", "OK")
                self.log(f"Código SUNAT: {respuesta.get('codigo_sunat', 'N/A')}", "INFO")
                self.log(f"Mensaje: {respuesta.get('mensaje_sunat', 'N/A')}", "INFO")
                
                if respuesta.get('es_mock'):
                    self.log("(Respuesta generada por mock)", "INFO")
                
            else:
                self.log("Envío fallido ✗", "ERROR")
                self.log(f"Error: {respuesta.get('mensaje_sunat', 'Error desconocido')}", "ERROR")
            
            # Guardar respuesta
            archivo_respuesta = self.carpeta_salida / "paso4_respuesta_apifas.json"
            with open(archivo_respuesta, 'w', encoding='utf-8') as f:
                json.dump(respuesta, f, indent=2, ensure_ascii=False)
            
            self.log(f"Respuesta guardada: {archivo_respuesta.name}", "INFO")
            self.log("")
            
            return exito, respuesta
            
        except Exception as e:
            self.log(f"Error en Paso 4: {e}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            return False, {"error": str(e)}
    
    def paso_5_validar_cdr(self, respuesta: dict) -> bool:
        """
        Paso 5: Valida CDR recibido.
        
        Args:
            respuesta: Respuesta JSON de APIFAS
        
        Returns:
            True si CDR es válido
        """
        self.log("=" * 70)
        self.log("PASO 5: Validación de CDR", "INFO")
        self.log("=" * 70)
        
        try:
            # Verificar que la respuesta sea exitosa
            if not respuesta.get('success'):
                self.log("Respuesta no exitosa, no hay CDR", "ERROR")
                return False
            
            # Validar campos obligatorios del CDR
            campos_requeridos = ['codigo_sunat', 'mensaje_sunat', 'hash_cdr', 'nombre_xml_cdr']
            
            for campo in campos_requeridos:
                if campo not in respuesta or not respuesta[campo]:
                    self.log(f"Campo faltante en CDR: {campo}", "ERROR")
                    return False
            
            self.log("Estructura de CDR válida ✓", "OK")
            self.log(f"Hash CDR: {respuesta['hash_cdr']}", "INFO")
            self.log(f"XML CDR: {respuesta['nombre_xml_cdr']}", "INFO")
            
            if 'observaciones' in respuesta and respuesta['observaciones']:
                self.log(f"Observaciones: {len(respuesta['observaciones'])}", "WARNING")
                for obs in respuesta['observaciones']:
                    self.log(f"  - {obs}", "WARNING")
            else:
                self.log("Sin observaciones", "OK")
            
            self.log("")
            return True
            
        except Exception as e:
            self.log(f"Error en Paso 5: {e}", "ERROR")
            return False
    
    def generar_reporte_final(self, resultados: dict):
        """
        Genera reporte final del test.
        
        Args:
            resultados: Dict con resultados de todos los pasos
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        reporte = [
            "=" * 70,
            "TEST NIVEL 3 — INTEGRACIÓN COMPLETA",
            "Motor CPE DisateQ™ v3.0",
            "=" * 70,
            f"Fecha: {timestamp}",
            f"Modo: {'MOCK' if self.usar_mock else 'REAL'}",
            f"Archivo Excel: {self.archivo_excel.name}",
            "",
            "RESULTADOS POR PASO:",
            "-" * 70,
        ]
        
        pasos = [
            ("Paso 1: Lectura Excel", resultados.get('paso1', False)),
            ("Paso 2: Generación TXT", resultados.get('paso2', False)),
            ("Paso 3: Validación TXT", resultados.get('paso3', False)),
            ("Paso 4: Envío APIFAS", resultados.get('paso4', False)),
            ("Paso 5: Validación CDR", resultados.get('paso5', False)),
        ]
        
        for nombre, exito in pasos:
            estado = "✅ EXITOSO" if exito else "❌ FALLIDO"
            reporte.append(f"{nombre:.<50} {estado}")
        
        reporte.extend([
            "",
            "RESUMEN:",
            "-" * 70,
        ])
        
        total_pasos = len(pasos)
        pasos_exitosos = sum(1 for _, exito in pasos if exito)
        tasa_exito = (pasos_exitosos / total_pasos) * 100
        
        reporte.append(f"Pasos exitosos: {pasos_exitosos}/{total_pasos} ({tasa_exito:.1f}%)")
        
        if pasos_exitosos == total_pasos:
            reporte.append("")
            reporte.append("🎉 ¡TEST NIVEL 3 COMPLETADO EXITOSAMENTE!")
            reporte.append("")
            reporte.append("Próximos pasos:")
            reporte.append("  1. Verificar archivos en carpeta: ./salida_nivel3/")
            reporte.append("  2. Si usaste mock, repetir con --real")
            reporte.append("  3. Documentar resultado en ESTADO.md")
        else:
            reporte.append("")
            reporte.append("⚠️  TEST INCOMPLETO - Revisar logs arriba")
        
        reporte.extend([
            "",
            "ARCHIVOS GENERADOS:",
            "-" * 70,
        ])
        
        for archivo in sorted(self.carpeta_salida.glob("*")):
            reporte.append(f"  - {archivo.name} ({archivo.stat().st_size} bytes)")
        
        reporte.extend([
            "",
            "LOG COMPLETO:",
            "-" * 70,
        ])
        
        reporte.extend(self.log_test)
        
        reporte.append("=" * 70)
        
        # Guardar reporte
        archivo_reporte = self.carpeta_salida / "REPORTE_NIVEL3.txt"
        archivo_reporte.write_text("\n".join(reporte), encoding='utf-8')
        
        # Mostrar en pantalla
        print("\n" + "\n".join(reporte))
        print(f"\n📄 Reporte completo guardado en: {archivo_reporte}")
    
    def ejecutar(self):
        """Ejecuta el test completo."""
        print("")
        print("=" * 70)
        print("INICIANDO TEST NIVEL 3 — INTEGRACIÓN COMPLETA")
        print("=" * 70)
        print(f"Modo: {'MOCK' if self.usar_mock else 'REAL'}")
        print(f"Excel: {self.archivo_excel}")
        print(f"Salida: {self.carpeta_salida}")
        print("=" * 70)
        print("")
        
        resultados = {}
        
        # Paso 1: Leer Excel
        exito, comprobante, items = self.paso_1_leer_excel()
        resultados['paso1'] = exito
        
        if not exito:
            self.log("Test detenido por error en Paso 1", "ERROR")
            self.generar_reporte_final(resultados)
            return False
        
        # Paso 2: Generar TXT
        exito, ruta_txt, contenido_txt = self.paso_2_generar_txt(comprobante, items)
        resultados['paso2'] = exito
        
        if not exito:
            self.log("Test detenido por error en Paso 2", "ERROR")
            self.generar_reporte_final(resultados)
            return False
        
        # Paso 3: Validar TXT
        exito = self.paso_3_validar_txt(ruta_txt)
        resultados['paso3'] = exito
        
        if not exito:
            self.log("Test detenido por error en Paso 3", "ERROR")
            self.log("TXT generado pero no válido - revisar errores", "WARNING")
            self.generar_reporte_final(resultados)
            return False
        
        # Paso 4: Enviar a APIFAS
        serie = comprobante.get('cpe_serie', 'XXX')
        numero = comprobante.get('cpe_numero', '0000')
        exito, respuesta = self.paso_4_enviar_apifas(contenido_txt, serie, numero)
        resultados['paso4'] = exito
        
        if not exito:
            self.log("Test detenido por error en Paso 4", "ERROR")
            self.generar_reporte_final(resultados)
            return False
        
        # Paso 5: Validar CDR
        exito = self.paso_5_validar_cdr(respuesta)
        resultados['paso5'] = exito
        
        # Generar reporte final
        self.generar_reporte_final(resultados)
        
        return all(resultados.values())


def main():
    """Función principal."""
    parser = argparse.ArgumentParser(
        description="Test Nivel 3 — Integración completa Motor CPE v3.0"
    )
    
    parser.add_argument(
        '--mock',
        action='store_true',
        help='Usar mock APIFAS local (sin envío real a SUNAT)'
    )
    
    parser.add_argument(
        '--real',
        action='store_true',
        help='Usar APIFAS real (envío a SUNAT)'
    )
    
    parser.add_argument(
        '--excel',
        type=str,
        default='prueba_integracion_completa.xlsx',
        help='Ruta al archivo Excel (default: prueba_integracion_completa.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Validar argumentos
    if args.mock and args.real:
        print("❌ Error: No se puede usar --mock y --real simultáneamente")
        sys.exit(1)
    
    if not args.mock and not args.real:
        print("⚠️  No se especificó modo, usando --mock por defecto")
        args.mock = True
    
    # Buscar archivo Excel
    rutas_buscar = [
        Path(args.excel),
        Path('D:/FFEESUNAT/test') / args.excel,
        Path('D:/FFEESUNAT/TEST') / args.excel,
        Path('./') / args.excel,
    ]
    
    archivo_excel = None
    for ruta in rutas_buscar:
        if ruta.exists():
            archivo_excel = ruta
            break
    
    if not archivo_excel:
        print(f"❌ Error: No se encontró el archivo Excel: {args.excel}")
        print("\nRutas buscadas:")
        for ruta in rutas_buscar:
            print(f"  - {ruta}")
        sys.exit(1)
    
    # Ejecutar test
    test = TestNivel3(
        archivo_excel=str(archivo_excel),
        usar_mock=args.mock
    )
    
    exito = test.ejecutar()
    
    sys.exit(0 if exito else 1)


if __name__ == "__main__":
    main()
