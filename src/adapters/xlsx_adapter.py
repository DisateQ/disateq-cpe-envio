"""
adapters/xlsx_adapter.py
========================
Adaptador Excel para DisateQ CPE™.
Lee la hoja _CPE exportada por DisateQ POS™.

Contrato de datos v1.1 — Abril 2026 (cerrado).
Campos segun catalogo definitivo _CPE v1.1.

Firma uniforme:
    XlsxAdapter().leer(ruta) -> list[(cabecera, items)]
"""

from __future__ import annotations

import logging
from pathlib import Path
from adapters.base_adapter import BaseAdapter, AdapterError

try:
    import openpyxl
except ImportError as e:
    raise ImportError("openpyxl no instalado. Ejecutar: pip install openpyxl") from e

log = logging.getLogger(__name__)

_HOJA = "_CPE"

# Columnas requeridas — contrato v1.1
_COLS_CABECERA = {
    "cpe_tipo",
    "cpe_serie",
    "cpe_numero",
    "cpe_fecha",
    "cli_tipo_doc",
    "cli_num_doc",
    "cli_nombre",
    "cli_direccion",
    "venta_subtotal",
    "venta_igv",
    "venta_total",
    "pago_forma",
    "estado",
}

_COLS_ITEM = {
    "item_orden",
    "item_descripcion",
    "item_unidad",
    "item_unspsc",
    "item_cantidad",
    "item_precio_unit",
    "item_valor_unitario",
    "item_subtotal_sin_igv",
    "item_igv",
    "item_total",
    "item_afectacion_igv",
}


# ── Helpers ───────────────────────────────────────────────────

def _str(valor, default="") -> str:
    if valor is None:
        return default
    return str(valor).strip() or default


def _float(valor, campo: str) -> float:
    try:
        return float(valor)
    except (TypeError, ValueError):
        raise AdapterError(f"Valor no numerico en columna '{campo}': {valor!r}")


def _int(valor, campo: str) -> int:
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        raise AdapterError(f"Valor entero invalido en columna '{campo}': {valor!r}")


def _fecha(valor, campo: str) -> str:
    """Normaliza fecha a DD-MM-YYYY."""
    from datetime import date, datetime
    if isinstance(valor, (date, datetime)):
        return valor.strftime("%d-%m-%Y")
    s = _str(valor)
    if not s:
        raise AdapterError(f"Fecha vacia en columna '{campo}'")
    # YYYY-MM-DD → DD-MM-YYYY
    if len(s) == 10 and s[4] == "-":
        a, m, d = s.split("-")
        return f"{d}-{m}-{a}"
    # DD-MM-YYYY
    if len(s) == 10 and s[2] == "-":
        return s
    raise AdapterError(f"Formato de fecha no reconocido en '{campo}': {s!r}")


def _validar_columnas(encabezados: set, requeridas: set, contexto: str):
    faltantes = requeridas - encabezados
    if faltantes:
        raise AdapterError(
            f"{contexto}: columnas requeridas ausentes en _CPE: "
            + ", ".join(sorted(faltantes))
        )


def _mapear_forma_pago(forma: str) -> str:
    """Mapea formas de pago POS → formato interno CPE."""
    mapa = {
        "EFECTIVO":      "Contado",
        "YAPE":          "Contado",
        "TRANSFERENCIA": "Contado",
        "TARJETA":       "Contado",
        "CREDITO":       "Credito",
    }
    return mapa.get(forma.upper().strip(), "Contado")


def _mapear_tipo_doc(cpe_tipo: str) -> str:
    """01=Factura, 03=Boleta."""
    return cpe_tipo.strip() if cpe_tipo.strip() in ("01", "03") else "03"


# ── Clase adaptadora ─────────────────────────────────────────

class XlsxAdapter(BaseAdapter):
    """
    Adaptador para leer comprobantes desde hoja _CPE de DisateQ POS™.
    Contrato de datos v1.1.
    """

    @property
    def nombre(self) -> str:
        return "XlsxAdapter (_CPE v1.1)"

    def leer(self, ruta: str) -> list:
        """
        Lee el archivo xlsx y retorna lista de (cabecera, items).
        Cada elemento es un comprobante listo para normalizar_desde_cpe().

        Args:
            ruta: path al archivo .xlsx (puede incluir rutas secundarias con |)

        Returns:
            list de (cabecera: dict, items: list[dict])

        Raises:
            AdapterError si el archivo, la hoja o los datos son invalidos
        """
        # Si vienen múltiples rutas separadas por |, usar la primera que tenga xlsx
        rutas = [r.strip() for r in ruta.split("|") if r.strip()]
        ruta_xlsx = None

        for r in rutas:
            p = Path(r)
            if p.suffix.lower() in (".xlsx", ".xlsm"):
                if p.exists():
                    ruta_xlsx = p
                    break
            # Si es carpeta, buscar _CPE.xlsx dentro
            if p.is_dir():
                candidatos = list(p.glob("_CPE*.xlsx")) + list(p.glob("_CPE*.xlsm"))
                if candidatos:
                    ruta_xlsx = candidatos[0]
                    break

        if not ruta_xlsx:
            raise AdapterError(
                f"Archivo _CPE.xlsx no encontrado en las rutas configuradas: {rutas}"
            )

        try:
            wb = openpyxl.load_workbook(str(ruta_xlsx), data_only=True)
        except Exception as e:
            raise AdapterError(f"No se pudo abrir el archivo xlsx: {e}") from e

        if _HOJA not in wb.sheetnames:
            raise AdapterError(
                f"Hoja '{_HOJA}' no encontrada en {ruta_xlsx.name}. "
                f"Hojas disponibles: {wb.sheetnames}"
            )

        ws = wb[_HOJA]
        filas = list(ws.iter_rows(values_only=True))

        if len(filas) < 2:
            raise AdapterError(
                f"Hoja _CPE sin datos — se esperan al menos fila 1=headers, fila 2+=datos"
            )

        # Fila 1: encabezados
        nombres = [_str(c).lower() for c in filas[0]]
        encabezados_set = set(nombres)

        _validar_columnas(encabezados_set, _COLS_CABECERA, "Cabecera")
        _validar_columnas(encabezados_set, _COLS_ITEM,     "Items")

        idx = {nombre: i for i, nombre in enumerate(nombres)}

        def col(fila, nombre):
            i = idx.get(nombre)
            return fila[i] if i is not None and i < len(fila) else None

        # Agrupar filas por comprobante (cpe_serie + cpe_numero)
        comprobantes: dict[tuple, dict] = {}

        for fila_n, fila in enumerate(filas[1:], start=2):
            if all(v is None for v in fila):
                continue

            # Solo procesar filas con estado BORRADOR o SIN CONEXION
            estado = _str(col(fila, "estado")).upper()
            if estado not in ("BORRADOR", "SIN CONEXION", ""):
                continue

            serie  = _str(col(fila, "cpe_serie"))
            numero = _str(col(fila, "cpe_numero"))
            clave  = (serie, numero)

            if clave not in comprobantes:
                # Primera fila de este comprobante — leer cabecera
                try:
                    cabecera = {
                        "tipo_doc":            _mapear_tipo_doc(_str(col(fila, "cpe_tipo"))),
                        "serie":               serie,
                        "numero":              _int(col(fila, "cpe_numero"), "cpe_numero"),
                        "fecha_emision":       _fecha(col(fila, "cpe_fecha"), "cpe_fecha"),
                        "moneda":              "PEN",
                        "forma_pago":          _mapear_forma_pago(_str(col(fila, "pago_forma"), "EFECTIVO")),
                        "cliente_tipo_doc":    _str(col(fila, "cli_tipo_doc"), "-"),
                        "cliente_numero_doc":  _str(col(fila, "cli_num_doc"), "00000000"),
                        "cliente_denominacion":_str(col(fila, "cli_nombre"), "CLIENTES VARIOS"),
                        "cliente_direccion":   _str(col(fila, "cli_direccion"), "-"),
                        "total_gravada":       0.0,
                        "total_exonerada":     0.0,
                        "total_inafecta":      0.0,
                        "total_igv":           _float(col(fila, "venta_igv"),    "venta_igv"),
                        "total_icbper":        0.0,
                        "total":               _float(col(fila, "venta_total"),  "venta_total"),
                    }
                except AdapterError:
                    raise
                except Exception as e:
                    raise AdapterError(f"Error leyendo cabecera en fila {fila_n}: {e}") from e

                comprobantes[clave] = {"cabecera": cabecera, "items": []}

            # Leer item de esta fila
            afectacion = _str(col(fila, "item_afectacion_igv"), "10")
            try:
                item = {
                    "codigo":           _str(col(fila, "item_codigo"), "000"),
                    "descripcion":      _str(col(fila, "item_descripcion"), "SIN DESCRIPCION").upper(),
                    "unspsc":           _str(col(fila, "item_unspsc"), "10000000"),
                    "unidad":           _str(col(fila, "item_unidad"), "NIU"),
                    "cantidad":         _float(col(fila, "item_cantidad"),        f"item_cantidad fila {fila_n}"),
                    "precio_con_igv":   _float(col(fila, "item_precio_unit"),     f"item_precio_unit fila {fila_n}"),
                    "precio_sin_igv":   _float(col(fila, "item_valor_unitario"),  f"item_valor_unitario fila {fila_n}"),
                    "subtotal_sin_igv": _float(col(fila, "item_subtotal_sin_igv"),f"item_subtotal_sin_igv fila {fila_n}"),
                    "igv":              _float(col(fila, "item_igv"),              f"item_igv fila {fila_n}"),
                    "total":            _float(col(fila, "item_total"),            f"item_total fila {fila_n}"),
                    "afectacion_igv":   afectacion,
                }
            except AdapterError:
                raise
            except Exception as e:
                raise AdapterError(f"Error leyendo item en fila {fila_n}: {e}") from e

            comprobantes[clave]["items"].append(item)

        if not comprobantes:
            raise AdapterError(
                "No se encontraron comprobantes en estado BORRADOR o SIN CONEXION en _CPE."
            )

        # Calcular totales por afectacion
        resultado = []
        for clave, datos in comprobantes.items():
            cab   = datos["cabecera"]
            items = datos["items"]

            gravada   = sum(i["subtotal_sin_igv"] for i in items if i["afectacion_igv"] == "10")
            exonerada = sum(i["total"]            for i in items if i["afectacion_igv"] == "20")
            inafecta  = sum(i["total"]            for i in items if i["afectacion_igv"] == "30")

            cab["total_gravada"]   = round(gravada,   8)
            cab["total_exonerada"] = round(exonerada, 8)
            cab["total_inafecta"]  = round(inafecta,  8)

            log.info(
                f"XlsxAdapter: {cab['serie']}-{str(cab['numero']).zfill(8)} "
                f"| {len(items)} item(s) | total S/ {cab['total']}"
            )
            resultado.append((cab, items))

        return resultado


# ── Función legacy (compatibilidad hacia atrás) ───────────────

def leer(ruta: str) -> tuple[dict, list[dict]]:
    """
    Función legacy — mantiene compatibilidad con código anterior.
    Retorna solo el primer comprobante encontrado.
    Para uso completo usar XlsxAdapter().leer(ruta).
    """
    adapter   = XlsxAdapter()
    resultado = adapter.leer(ruta)
    if not resultado:
        raise AdapterError("No se encontraron comprobantes en _CPE.")
    cabecera, items = resultado[0]
    return cabecera, items