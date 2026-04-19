"""
source_explorer.py
==================
Explorador universal de fuentes de datos — Motor CPE DisateQ™ v3.0

PROPÓSITO:
  Herramienta de inspección técnica para identificar campos disponibles
  en cualquier fuente de datos (DBF, XLSX, CSV, MDB, ODBC, SQL Server, etc.)
  ANTES de mapear hacia el contrato _CPE.

FLUJO:
  1. Técnico ejecuta: python tools/source_explorer.py --source facturacion.dbf
  2. El explorador detecta tipo, lee estructura y muestra campos disponibles
  3. Técnico usa esa información para construir el YAML de mapeo
  4. El adapter usa ese YAML para transformar source → _CPE

USO:
  python tools/source_explorer.py --source C:\Sistemas\ventas.dbf
  python tools/source_explorer.py --source ventas.xlsx --verbose
  python tools/source_explorer.py --source VENTAS --connection "Driver={SQL Server};..."
"""

import sys
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
from enum import Enum


class SourceType(Enum):
    """Tipos de fuentes de datos soportadas."""
    DBF = "dbf"
    XLSX = "xlsx"
    CSV = "csv"
    MDB = "mdb"
    ODBC = "odbc"
    SQLSERVER = "sqlserver"
    POSTGRESQL = "postgresql"
    MYSQL = "mysql"
    ORACLE = "oracle"
    UNKNOWN = "unknown"


@dataclass
class FieldInfo:
    """Información de un campo detectado."""
    name: str
    type: str
    length: Optional[int] = None
    decimals: Optional[int] = None
    nullable: bool = True
    sample_values: List[Any] = None
    
    def __post_init__(self):
        if self.sample_values is None:
            self.sample_values = []


@dataclass
class TableInfo:
    """Información de una tabla/hoja detectada."""
    name: str
    record_count: int
    fields: List[FieldInfo]
    source_type: SourceType


class SourceExplorer:
    """
    Explorador universal de fuentes de datos.
    Detecta tipo, lee estructura y extrae muestras.
    """
    
    def __init__(self, source_path: str, connection_string: Optional[str] = None):
        self.source_path = source_path
        self.connection_string = connection_string
        self.source_type = self._detect_type()
    
    def _detect_type(self) -> SourceType:
        """Detecta el tipo de fuente por extensión o connection string."""
        if self.connection_string:
            cs_lower = self.connection_string.lower()
            if 'driver={sql server' in cs_lower or 'sqlserver' in cs_lower:
                return SourceType.SQLSERVER
            elif 'postgresql' in cs_lower or 'postgres' in cs_lower:
                return SourceType.POSTGRESQL
            elif 'mysql' in cs_lower:
                return SourceType.MYSQL
            elif 'oracle' in cs_lower:
                return SourceType.ORACLE
            else:
                return SourceType.ODBC
        
        path = Path(self.source_path)
        ext = path.suffix.lower()
        
        type_map = {
            '.dbf': SourceType.DBF,
            '.xlsx': SourceType.XLSX,
            '.xls': SourceType.XLSX,
            '.csv': SourceType.CSV,
            '.mdb': SourceType.MDB,
            '.accdb': SourceType.MDB,
        }
        
        return type_map.get(ext, SourceType.UNKNOWN)
    
    def explore(self, max_samples: int = 5) -> List[TableInfo]:
        """Explora la fuente y retorna información de tablas/hojas."""
        explorer_methods = {
            SourceType.DBF: self._explore_dbf,
            SourceType.XLSX: self._explore_xlsx,
            SourceType.CSV: self._explore_csv,
            SourceType.MDB: self._explore_mdb,
            SourceType.ODBC: self._explore_sql,
            SourceType.SQLSERVER: self._explore_sql,
            SourceType.POSTGRESQL: self._explore_sql,
            SourceType.MYSQL: self._explore_sql,
            SourceType.ORACLE: self._explore_sql,
        }
        
        method = explorer_methods.get(self.source_type)
        if not method:
            raise ValueError(f"Tipo de fuente no soportado: {self.source_type}")
        
        return method(max_samples)
    
    def _explore_dbf(self, max_samples: int) -> List[TableInfo]:
        """Explora archivo DBF usando dbfread."""
        try:
            from dbfread import DBF
        except ImportError:
            raise RuntimeError("dbfread no instalado. Ejecutar: pip install dbfread")
        
        path = Path(self.source_path)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.source_path}")
        
        table = DBF(str(path), encoding='latin-1', ignore_missing_memofile=True)
        
        fields = []
        for field in table.fields:
            field_info = FieldInfo(
                name=field.name,
                type=field.type,
                length=field.length,
                decimals=getattr(field, 'decimal_count', None),
                sample_values=[]
            )
            fields.append(field_info)
        
        records = list(table)
        record_count = len(records)
        
        if records:
            for record in records[:max_samples]:
                for field_info in fields:
                    value = record.get(field_info.name)
                    if value is not None and value != '':
                        field_info.sample_values.append(value)
        
        return [TableInfo(
            name=path.stem,
            record_count=record_count,
            fields=fields,
            source_type=SourceType.DBF
        )]
    
    def _explore_xlsx(self, max_samples: int) -> List[TableInfo]:
        """Explora archivo Excel usando openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl no instalado. Ejecutar: pip install openpyxl")
        
        path = Path(self.source_path)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.source_path}")
        
        wb = load_workbook(str(path), read_only=True, data_only=True)
        tables = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else f"Col{cell.column}")
            
            record_count = sheet.max_row - 1
            
            fields = []
            for idx, header in enumerate(headers, start=1):
                field_info = FieldInfo(name=header, type="text", sample_values=[])
                
                for row_num in range(2, min(2 + max_samples, sheet.max_row + 1)):
                    cell = sheet.cell(row=row_num, column=idx)
                    if cell.value is not None:
                        field_info.sample_values.append(cell.value)
                
                fields.append(field_info)
            
            tables.append(TableInfo(
                name=sheet_name,
                record_count=record_count,
                fields=fields,
                source_type=SourceType.XLSX
            ))
        
        wb.close()
        return tables
    
    def _explore_csv(self, max_samples: int) -> List[TableInfo]:
        """Explora archivo CSV."""
        import csv
        
        path = Path(self.source_path)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.source_path}")
        
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            
            fields = [FieldInfo(name=h, type="text", sample_values=[]) for h in headers]
            
            record_count = 0
            for row in reader:
                record_count += 1
                if record_count <= max_samples:
                    for field_info in fields:
                        value = row.get(field_info.name)
                        if value:
                            field_info.sample_values.append(value)
            
            return [TableInfo(
                name=path.stem,
                record_count=record_count,
                fields=fields,
                source_type=SourceType.CSV
            )]
    
    def _explore_mdb(self, max_samples: int) -> List[TableInfo]:
        """Explora base de datos Access."""
        try:
            import pyodbc
        except ImportError:
            raise RuntimeError("pyodbc no instalado. Ejecutar: pip install pyodbc")
        
        path = Path(self.source_path)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.source_path}")
        
        conn_str = f'Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={path};'
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        tables = []
        for table_info in cursor.tables(tableType='TABLE'):
            table_name = table_info.table_name
            
            fields = []
            cursor.execute(f"SELECT TOP 1 * FROM [{table_name}]")
            for column in cursor.description:
                field_info = FieldInfo(
                    name=column[0],
                    type=self._sql_type_to_string(column[1]),
                    length=column[2],
                    sample_values=[]
                )
                fields.append(field_info)
            
            cursor.execute(f"SELECT COUNT(*) FROM [{table_name}]")
            record_count = cursor.fetchone()[0]
            
            cursor.execute(f"SELECT TOP {max_samples} * FROM [{table_name}]")
            for row in cursor.fetchall():
                for idx, field_info in enumerate(fields):
                    value = row[idx]
                    if value is not None:
                        field_info.sample_values.append(value)
            
            tables.append(TableInfo(
                name=table_name,
                record_count=record_count,
                fields=fields,
                source_type=SourceType.MDB
            ))
        
        conn.close()
        return tables
    
    def _explore_sql(self, max_samples: int) -> List[TableInfo]:
        """Explora fuente SQL (universal para todos los SQL)."""
        try:
            import pyodbc
        except ImportError:
            raise RuntimeError("pyodbc no instalado. Ejecutar: pip install pyodbc")
        
        conn = pyodbc.connect(self.connection_string)
        cursor = conn.cursor()
        
        table_name = self.source_path
        
        fields = []
        cursor.execute(f"SELECT TOP 1 * FROM {table_name}")
        for column in cursor.description:
            field_info = FieldInfo(
                name=column[0],
                type=self._sql_type_to_string(column[1]),
                length=column[2],
                sample_values=[]
            )
            fields.append(field_info)
        
        cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
        record_count = cursor.fetchone()[0]
        
        cursor.execute(f"SELECT TOP {max_samples} * FROM {table_name}")
        for row in cursor.fetchall():
            for idx, field_info in enumerate(fields):
                value = row[idx]
                if value is not None:
                    field_info.sample_values.append(value)
        
        conn.close()
        
        return [TableInfo(
            name=table_name,
            record_count=record_count,
            fields=fields,
            source_type=self.source_type
        )]
    
    def _sql_type_to_string(self, type_code) -> str:
        """Convierte código de tipo SQL a string legible."""
        try:
            import pyodbc
            type_map = {
                pyodbc.SQL_CHAR: "char",
                pyodbc.SQL_VARCHAR: "varchar",
                pyodbc.SQL_LONGVARCHAR: "text",
                pyodbc.SQL_WCHAR: "nchar",
                pyodbc.SQL_WVARCHAR: "nvarchar",
                pyodbc.SQL_DECIMAL: "decimal",
                pyodbc.SQL_NUMERIC: "numeric",
                pyodbc.SQL_SMALLINT: "smallint",
                pyodbc.SQL_INTEGER: "int",
                pyodbc.SQL_REAL: "real",
                pyodbc.SQL_FLOAT: "float",
                pyodbc.SQL_DOUBLE: "float",
                pyodbc.SQL_BIT: "bit",
                pyodbc.SQL_TINYINT: "tinyint",
                pyodbc.SQL_BIGINT: "bigint",
                pyodbc.SQL_TYPE_DATE: "date",
                pyodbc.SQL_TYPE_TIME: "time",
                pyodbc.SQL_TYPE_TIMESTAMP: "datetime",
            }
            return type_map.get(type_code, f"unknown({type_code})")
        except:
            return "unknown"


def format_report(tables: List[TableInfo], verbose: bool = False) -> str:
    """Genera reporte legible de la estructura detectada."""
    lines = []
    lines.append("=" * 80)
    lines.append("  SOURCE EXPLORER — Motor CPE DisateQ™ v3.0")
    lines.append("  Estructura detectada de la fuente de datos")
    lines.append("=" * 80)
    lines.append("")
    
    for table in tables:
        lines.append(f"TABLA/HOJA: {table.name}")
        lines.append(f"Tipo fuente: {table.source_type.value.upper()}")
        lines.append(f"Registros:   {table.record_count:,}")
        lines.append("-" * 80)
        lines.append("")
        
        lines.append(f"{'CAMPO':<30} {'TIPO':<15} {'LARGO':<10} {'DEC':<5}")
        lines.append("-" * 80)
        
        for field in table.fields:
            length_str = str(field.length) if field.length else "-"
            dec_str = str(field.decimals) if field.decimals else "-"
            lines.append(f"{field.name:<30} {field.type:<15} {length_str:<10} {dec_str:<5}")
            
            if verbose and field.sample_values:
                samples_str = ", ".join(str(v)[:30] for v in field.sample_values[:3])
                lines.append(f"  └─ Ejemplos: {samples_str}")
                lines.append("")
        
        lines.append("")
        lines.append("=" * 80)
        lines.append("")
    
    total_fields = sum(len(t.fields) for t in tables)
    lines.append(f"RESUMEN:")
    lines.append(f"  Tablas/hojas detectadas: {len(tables)}")
    lines.append(f"  Campos totales:          {total_fields}")
    lines.append("")
    lines.append("PRÓXIMO PASO:")
    lines.append("  Crear archivo YAML de mapeo con estos nombres de campo")
    lines.append("  Ver: docs/mapping_examples/ejemplo_dbf.yaml")
    lines.append("=" * 80)
    
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Source Explorer — Motor CPE DisateQ™ v3.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:

  # Explorar archivo DBF
  python tools/source_explorer.py --source facturacion.dbf

  # Explorar Excel con valores de ejemplo
  python tools/source_explorer.py --source ventas.xlsx --verbose

  # Explorar tabla SQL Server
  python tools/source_explorer.py --source VENTAS \
    --connection "Driver={SQL Server};Server=localhost;Database=ERP"

  # Guardar reporte en archivo
  python tools/source_explorer.py --source datos.dbf --output reporte.txt
        """
    )
    
    parser.add_argument('--source', required=True, help='Ruta al archivo o nombre de tabla')
    parser.add_argument('--connection', help='Connection string para SQL (opcional)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Mostrar valores de ejemplo')
    parser.add_argument('--samples', type=int, default=5, help='Número de valores de ejemplo (default: 5)')
    parser.add_argument('--output', '-o', help='Guardar reporte en archivo')
    
    args = parser.parse_args()
    
    try:
        explorer = SourceExplorer(args.source, args.connection)
        
        print(f"Explorando: {args.source}")
        print(f"Tipo detectado: {explorer.source_type.value.upper()}")
        print("Leyendo estructura...\n")
        
        tables = explorer.explore(max_samples=args.samples)
        report = format_report(tables, verbose=args.verbose)
        
        if args.output:
            Path(args.output).write_text(report, encoding='utf-8')
            print(f"✓ Reporte guardado en: {args.output}")
        else:
            print(report)
        
        return 0
        
    except Exception as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
